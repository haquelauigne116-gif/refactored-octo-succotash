"""
server.py — FastAPI 路由入口 (精简版，业务逻辑已拆分到各模块)
"""
from contextlib import asynccontextmanager
from typing import Optional

import asyncio
import os
import threading
import uuid as _uuid

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, Form, Query  # type: ignore[import]
from fastapi.middleware.cors import CORSMiddleware  # type: ignore[import]
from fastapi.staticfiles import StaticFiles  # type: ignore[import]
from fastapi.responses import FileResponse, StreamingResponse  # type: ignore[import]
from pydantic import BaseModel  # type: ignore[import]

# 修复 Windows 下 ES Module 的 MIME 类型问题
import mimetypes
mimetypes.add_type('application/javascript', '.js')
mimetypes.add_type('text/css', '.css')

from backend.config import (  # type: ignore[import]
    SYSTEM_PROMPT, INDEX_HTML, FRONTEND_DIR,
    API_PROVIDERS, APP_SETTINGS, SESSION_DIR,
    get_client, get_model_caps, load_settings, save_settings, save_providers_config,
    MEMORY_FILE,
    load_notification_config, save_notification_config,
)
from backend.ai.session_manager import SessionManager  # type: ignore[import]
from backend.ai.rag_engine import RAGEngine  # type: ignore[import]
from backend.scheduling.task_scheduler import TaskScheduler  # type: ignore[import]
from backend.integrations.messaging import (  # type: ignore[import]
    NotificationManager, WebSocketChannel, QQChannel,
    NapCatWatchdog,
)
from backend.integrations.messaging import QQChatHandler  # type: ignore[import]
from backend.ai.memory_worker import DailyMemoryWorker  # type: ignore[import]
from backend.file_storage import MinIOManager  # type: ignore[import]
from backend.integrations.mcp import mcp_mgr  # type: ignore[import]
from backend.scheduling.schedule_manager import ScheduleManager  # type: ignore[import]


# ====== 初始化核心组件 ======
session_mgr = SessionManager()
rag = RAGEngine()
task_scheduler = TaskScheduler()
minio_mgr = MinIOManager()
schedule_mgr = ScheduleManager()

# 当前对话模型状态 (从配置中读取持久化状态)
current_provider_id = APP_SETTINGS.get("chat_provider", "deepseek")
current_model = APP_SETTINGS.get("chat_model", "deepseek-chat")

if API_PROVIDERS:
    if current_provider_id not in API_PROVIDERS:
        current_provider_id = list(API_PROVIDERS.keys())[0]
        current_model = API_PROVIDERS[current_provider_id]["models"][0]["id"]

client = get_client(current_provider_id)

# ====== WebSocket 连接管理 ======
ws_clients: set[WebSocket] = set()
_event_loop = None  # 主事件循环引用
notification_mgr: NotificationManager | None = None  # 多通道通知管理器
qq_chat_handler: QQChatHandler | None = None  # QQ 聊天处理器（供 /ws/qq 使用）
napcat_watchdog: NapCatWatchdog | None = None  # NapCat 连接/登录守护

async def broadcast(data: dict):
    """向所有连接的 WebSocket 客户端广播消息"""
    dead: list[WebSocket] = []
    for client in list(ws_clients):
        try:
            await client.send_json(data)
        except Exception:
            dead.append(client)
    for c in dead:
        ws_clients.discard(c)

def broadcast_sync(data: dict):
    """同步回调桥接：从 APScheduler 线程安全地推送到 async WebSocket"""
    if _event_loop and _event_loop.is_running():
        asyncio.run_coroutine_threadsafe(broadcast(data), _event_loop)

# ====== FastAPI 应用 ======

@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期：启动时开启调度器和通知通道，关闭时停止"""
    global _event_loop, notification_mgr, qq_chat_handler, napcat_watchdog
    _event_loop = asyncio.get_running_loop()

    # 初始化多通道通知管理器
    notif_config = load_notification_config()
    channels_cfg = notif_config.get("channels", {})

    notification_mgr = NotificationManager()

    # WebSocket 通道
    ws_channel = WebSocketChannel()
    ws_cfg = channels_cfg.get("websocket", {})
    ws_channel.enabled = ws_cfg.get("enabled", True)
    ws_channel.set_broadcast_fn(broadcast_sync)
    notification_mgr.add_channel(ws_channel)



    # QQ 通道 (NapCat / OneBot 11)
    qq_cfg = channels_cfg.get("qq", {})
    qq_channel = QQChannel(
        napcat_http_url=qq_cfg.get("napcat_http_url", "http://127.0.0.1:3000"),
        napcat_token=qq_cfg.get("napcat_token", ""),
        msg_type=qq_cfg.get("msg_type", "private"),
        target_user_ids=qq_cfg.get("target_user_ids", []),
        target_group_ids=qq_cfg.get("target_group_ids", []),
        enabled=qq_cfg.get("enabled", False),
    )
    notification_mgr.add_channel(qq_channel)

    # 创建 QQ 聊天处理器并注入
    qq_chat_handler = QQChatHandler(rag=rag, minio_mgr=minio_mgr)
    qq_chat_handler.set_task_scheduler(task_scheduler)
    qq_chat_handler.set_schedule_mgr(schedule_mgr)
    qq_chat_handler.set_qq_channel(qq_channel)

    # NapCat 守护进程 — 监控 WebSocket 心跳 + QQ 登录状态
    def _watchdog_alert(level: str, message: str):
        """守护告警 → WebSocket 广播 + 日志"""
        broadcast_sync({"type": "napcat_alert", "level": level, "message": message})

    napcat_watchdog = NapCatWatchdog(
        napcat_http_url=qq_cfg.get("napcat_http_url", "http://127.0.0.1:3000"),
        napcat_token=qq_cfg.get("napcat_token", ""),
        heartbeat_timeout=90,
        login_check_interval=60,
        napcat_cmd=r"D:\NapCat\NapCat.44498.Shell\NapCatWinBootMain.exe 3920800540",
        max_restart_attempts=5,
        restart_cooldown=600,
        on_alert=_watchdog_alert,
    )
    if qq_cfg.get("enabled", False):
        napcat_watchdog.start()
        print("[System] NapCat 守护进程已启动 (心跳监控 + 登录检测)")

    # 启动所有通道
    notification_mgr.start()

    # 注入到任务调度器
    task_scheduler.notification_manager = notification_mgr
    task_scheduler.mcp_manager = mcp_mgr  # 注入 MCP 管理器，使定时任务支持工具调用
    task_scheduler.start()

    # 注入日程管理器
    schedule_mgr.task_scheduler = task_scheduler
    schedule_mgr.notification_manager = notification_mgr

    from apscheduler.triggers.cron import CronTrigger  # type: ignore[import]
    from apscheduler.triggers.interval import IntervalTrigger  # type: ignore[import]

    # 注册每日记忆整理 (凌晨 2:00)
    memory_worker = DailyMemoryWorker()
    task_scheduler.register_system_job(
        job_id="daily_memory_extraction",
        func=memory_worker.run,
        trigger=CronTrigger(hour=2, minute=0),
    )
    print("[System] 每日记忆整理已注册 (凌晨 2:00)")

    # 注册日程提醒轮询 (每分钟扫描一次, 仅占 1 个 job)
    task_scheduler.register_system_job(
        job_id="schedule_reminder_poll",
        func=schedule_mgr.check_and_fire_reminders,
        trigger=IntervalTrigger(minutes=1),
    )
    print("[System] 日程提醒轮询已注册 (每分钟)")

    # 注册每日日程简报 (早上 8:00, 仅占 1 个 job)
    task_scheduler.register_system_job(
        job_id="daily_schedule_briefing",
        func=schedule_mgr.send_daily_briefing,
        trigger=CronTrigger(hour=8, minute=0),
    )
    print("[System] 每日日程简报已注册 (早上 8:00)")

    yield

    task_scheduler.shutdown()
    notification_mgr.shutdown()
    if napcat_watchdog:
        napcat_watchdog.stop()
    _event_loop = None

app = FastAPI(lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    max_age=3600,
)

# ====== 页面路由 ======

@app.get("/")
def serve_index():
    return FileResponse(INDEX_HTML)

# ====== 模型提供商路由 ======

@app.get("/providers")
def get_providers():
    return {
        "current_provider": current_provider_id,
        "current_model": current_model,
        "providers": [
            {"id": k, "name": v["name"], "models": v["models"]}
            for k, v in API_PROVIDERS.items()
        ],
    }

class ProviderRequest(BaseModel):
    provider_id: str
    model_id: str

@app.post("/switch_provider")
def switch_provider(request: ProviderRequest):
    global current_provider_id, client, current_model
    if request.provider_id not in API_PROVIDERS:
        return {"status": "error", "message": "未知的提供商"}
    models = [m["id"] for m in API_PROVIDERS[request.provider_id]["models"]]
    if request.model_id not in models:
        return {"status": "error", "message": "未知的模型"}
    current_provider_id = request.provider_id
    current_model = request.model_id
    client = get_client(current_provider_id)
    
    # 持久化主聊天模型设置
    APP_SETTINGS["chat_provider"] = current_provider_id
    APP_SETTINGS["chat_model"] = current_model
    save_settings(APP_SETTINGS)

    return {"status": "ok", "provider": current_provider_id, "model": current_model}

# ====== 设置路由 ======

@app.get("/api/settings")
def get_system_settings():
    safe_providers = {}
    for pid, pdata in API_PROVIDERS.items():
        key = pdata.get("api_key", "")
        masked = f"{key[:6]}...{key[-4:]}" if len(key) > 15 else "未配置或过短"
        safe_providers[pid] = {"name": pdata["name"], "api_key_masked": masked}
    return {"providers_status": safe_providers, "settings": APP_SETTINGS}

class SettingsUpdateRequest(BaseModel):
    api_keys: dict
    chat_provider: str = ""
    chat_model: str = ""
    summary_provider: str
    summary_model: str
    judge_provider: str
    judge_model: str
    file_provider: str
    file_model: str
    file_vision_provider: str = ""
    file_vision_model: str = ""
    task_provider: str
    task_model: str
    bailian_api_key: str = ""
    enable_mcp_for_chat: bool = False
    max_tool_loops: int = 6

@app.post("/api/settings")
def update_system_settings(req: SettingsUpdateRequest):
    global API_PROVIDERS
    # 更新 API Keys
    changed = False
    for pid, new_key in req.api_keys.items():
        if pid in API_PROVIDERS and new_key and len(new_key.strip()) > 10:
            API_PROVIDERS[pid]["api_key"] = new_key.strip()
            changed = True
    if changed:
        save_providers_config(API_PROVIDERS)

    # 更新系统设置
    if req.chat_provider and req.chat_model:
        APP_SETTINGS["chat_provider"] = req.chat_provider
        APP_SETTINGS["chat_model"] = req.chat_model
        global current_provider_id, current_model, client
        current_provider_id = req.chat_provider
        current_model = req.chat_model
        try:
            client = get_client(current_provider_id)
        except Exception:
            pass
    APP_SETTINGS["summary_provider"] = req.summary_provider
    APP_SETTINGS["summary_model"] = req.summary_model
    APP_SETTINGS["judge_provider"] = req.judge_provider
    APP_SETTINGS["judge_model"] = req.judge_model
    APP_SETTINGS["file_provider"] = req.file_provider
    APP_SETTINGS["file_model"] = req.file_model
    if req.file_vision_provider:
        APP_SETTINGS["file_vision_provider"] = req.file_vision_provider
    if req.file_vision_model:
        APP_SETTINGS["file_vision_model"] = req.file_vision_model
    APP_SETTINGS["task_provider"] = req.task_provider
    APP_SETTINGS["task_model"] = req.task_model
    if req.bailian_api_key:
        APP_SETTINGS["bailian_api_key"] = req.bailian_api_key
    APP_SETTINGS["enable_mcp_for_chat"] = req.enable_mcp_for_chat
    APP_SETTINGS["max_tool_loops"] = max(1, min(req.max_tool_loops, 20))  # 限制范围 1-20
    save_settings(APP_SETTINGS)
    return {"status": "ok"}

# ====== 会话路由 ======

@app.get("/sessions")
def list_sessions():
    return {"sessions": session_mgr.list_all()}

@app.post("/new_session")
def new_session():
    session_mgr.create_new()
    return {"status": "ok", "session_id": session_mgr.session_id}

class SwitchRequest(BaseModel):
    filename: str

@app.post("/switch_session")
def switch_session(request: SwitchRequest):
    result = session_mgr.switch_to(request.filename)
    if result is None:
        return {"status": "error", "message": "会话不存在"}
    return {"status": "ok", "messages": result["messages"], "events": result["events"], "render_events": result.get("render_events", [])}

@app.post("/delete_session")
def delete_session(request: SwitchRequest):
    ok = session_mgr.delete_session(request.filename)
    if not ok:
        return {"status": "error", "message": "会话不存在"}
    return {"status": "ok"}


# ====== 聊天路由 ======

class ChatRequest(BaseModel):
    user_word: str = ""
    attachments: list[dict] | None = None  # [{data: base64, mime: str, name: str}]

# ====== 附件文本提取 ======

def _extract_doc_text(att: dict) -> str:
    """从 base64 编码的附件提取可读文本 (PDF / txt / doc / docx / xls / xlsx 等)"""
    import base64 as _b64
    import io as _io
    import re as _re

    mime = att.get("mime", "")
    name = att.get("name", "")
    data_b64 = att.get("data", "")
    if not data_b64:
        return ""

    try:
        raw = _b64.b64decode(data_b64)
    except Exception:
        return ""

    lower_name = name.lower()

    # PDF: 使用 PyPDF2 提取
    if mime == "application/pdf" or lower_name.endswith(".pdf"):
        try:
            from PyPDF2 import PdfReader  # type: ignore[import]
            reader = PdfReader(_io.BytesIO(raw))
            pages: list[str] = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text.strip())
            full_text = "\n\n".join(pages)
            return full_text[:5000]  # type: ignore[index]
        except ImportError:
            return "[系统提示: 需要安装 PyPDF2 以提取 PDF 内容: pip install PyPDF2]"
        except Exception as e:
            return f"[PDF 解析失败: {e}]"

    # Word .docx (新格式)
    if lower_name.endswith(".docx"):
        try:
            from docx import Document  # type: ignore[import]
            doc = Document(_io.BytesIO(raw))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return text[:5000]  # type: ignore[index]
        except ImportError:
            return "[系统提示: 需要安装 python-docx 以提取 Word 内容: pip install python-docx]"
        except Exception as e:
            return f"[Word 解析失败: {e}]"

    # Word .doc (旧二进制格式)
    if lower_name.endswith(".doc") or mime == "application/msword":
        try:
            # 方法 1: 尝试 olefile 提取 Word 文档流
            try:
                import olefile  # type: ignore[import]
                ole = olefile.OleFileIO(_io.BytesIO(raw))
                if ole.exists("WordDocument"):
                    # 提取主文档流中的文本 (简易方式)
                    stream = ole.openstream("WordDocument").read()
                    # 去除 null 字节, 提取可读文本
                    text = stream.replace(b'\x00', b'').decode('utf-8', errors='ignore')
                    text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
                    # 过滤掉太短的片段 (二进制噪声)
                    parts = [p.strip() for p in text.split('\r') if len(p.strip()) > 2]
                    result = "\n".join(parts)
                    if len(result) > 50:
                        return result[:5000]  # type: ignore[index]
                ole.close()
            except ImportError:
                pass

            # 方法 2: 直接从原始二进制中提取文本 (通用回退)
            text = raw.replace(b'\x00', b'').decode('utf-8', errors='ignore')
            text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
            # 提取较长的连续文本段落
            parts = [p.strip() for p in _re.split(r'[\r\n]+', text) if len(p.strip()) > 3]
            result = "\n".join(parts)
            if len(result) > 50:
                return result[:5000]  # type: ignore[index]
            return "[.doc 文件文本提取量不足, 建议转换为 .docx 格式后重试]"
        except Exception as e:
            return f"[.doc 解析失败: {e}，建议转换为 .docx 格式]"

    # Excel .xlsx
    if lower_name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook  # type: ignore[import]
            wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
            rows_text: list[str] = []
            for ws in wb.worksheets:
                rows_text.append(f"=== 工作表: {ws.title} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows_text.append("\t".join(cells))
            wb.close()
            result = "\n".join(rows_text)
            return result[:5000]  # type: ignore[index]
        except ImportError:
            return "[系统提示: 需要安装 openpyxl 以提取 Excel 内容: pip install openpyxl]"
        except Exception as e:
            return f"[Excel 解析失败: {e}]"

    # Excel .xls (旧格式)
    if lower_name.endswith(".xls"):
        try:
            import xlrd  # type: ignore[import]
            book = xlrd.open_workbook(file_contents=raw)
            rows_text_xls: list[str] = []
            for sheet in book.sheets():
                rows_text_xls.append(f"=== 工作表: {sheet.name} ===")
                for rx in range(sheet.nrows):
                    cells = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
                    if any(cells):
                        rows_text_xls.append("\t".join(cells))
            result = "\n".join(rows_text_xls)
            return result[:5000]  # type: ignore[index]
        except ImportError:
            return "[系统提示: 需要安装 xlrd 以提取 .xls 内容: pip install xlrd]"
        except Exception as e:
            return f"[.xls 解析失败: {e}]"

    # 纯文本类：txt, md, csv, json 等
    text_types = ("text/", "application/json", "application/xml", "application/csv")
    text_exts = (".txt", ".md", ".csv", ".json", ".xml", ".log", ".yml", ".yaml", ".ini", ".cfg")
    if any(mime.startswith(t) for t in text_types) or any(lower_name.endswith(e) for e in text_exts):
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text = raw.decode("gbk")
            except Exception:
                return "[文件编码无法识别]"
        return text[:5000]  # type: ignore[index]

    return ""

@app.post("/chat")
def chat_with_ai(request: ChatRequest):
    """SSE 流式聊天端点"""
    user_text = request.user_word
    attachments = request.attachments or []

    # --- 分析附件类型 ---
    image_attachments = [a for a in attachments if a.get("mime", "").startswith("image/")]
    doc_attachments = [a for a in attachments if not a.get("mime", "").startswith("image/")]

    # --- 提取文档内容 (PDF/txt 等) ---
    doc_text_parts: list[str] = []
    for att in doc_attachments:
        print(f"[Chat] 尝试提取文档: name={att.get('name')}, mime={att.get('mime')}, data_len={len(att.get('data', ''))}")
        extracted = _extract_doc_text(att)
        if extracted:
            doc_text_parts.append(f"[文件: {att.get('name', '未知')}]\n{extracted}")
            print(f"[Chat] 文档提取成功: {att.get('name')}, 提取长度={len(extracted)}")
        else:
            print(f"[Chat] 文档提取为空: {att.get('name')}")

    # --- 统一意图分析（仅用于 task/schedule/MCP 检测，RAG/文件搜索移入循环） ---
    intent = rag.analyze_intent(session_mgr.messages, user_text)
    task_intent = intent["task_intent"]
    schedule_intent = intent.get("schedule_intent")

    # --- 第一条消息时初始化文件 ---
    if session_mgr.is_first_message:
        session_mgr.initialize_file()

    # --- 写入用户消息 ---
    attach_label = ""
    if attachments:
        names = [a.get("name", "文件") for a in attachments]
        attach_label = f" [附件: {', '.join(names)}]"
    session_mgr.append_user_message(user_text + attach_label)
    user_msg_index = len(session_mgr.messages) - 1  # 记录用户消息位置, 用于附件事件定位

    # --- 检查模型视觉能力 ---
    model_caps = get_model_caps(current_provider_id, current_model)
    has_vision = "vision" in model_caps

    # --- 构建 AI 消息列表（清理 HTML 防止模型模仿工具卡片） ---
    ai_messages = []
    for m in session_mgr.messages:
        if m.get("role") not in ("system", "user", "assistant"):
            continue
        if m.get("role") == "assistant":
            cleaned = SessionManager._strip_html_for_chat(m.get("content", ""))
            if cleaned:
                ai_messages.append({"role": "assistant", "content": cleaned})
        else:
            ai_messages.append(m)

    # 图片 + vision 模型 → Vision 格式
    if image_attachments and has_vision:
        last_user_idx = None
        for i in range(len(ai_messages) - 1, -1, -1):
            if ai_messages[i]["role"] == "user":
                last_user_idx = i
                break
        if last_user_idx is not None:
            content_parts: list[dict] = []
            if user_text:
                content_parts.append({"type": "text", "text": user_text})
            for img in image_attachments:
                data_url = f"data:{img['mime']};base64,{img['data']}"
                content_parts.append({"type": "image_url", "image_url": {"url": data_url}})
            if doc_text_parts:
                content_parts.append({"type": "text", "text": "\n\n".join(doc_text_parts)})
            ai_messages[last_user_idx] = {"role": "user", "content": content_parts}  # type: ignore[index]
    elif image_attachments and not has_vision:
        ai_messages.insert(-1, {"role": "system", "content": "用户发送了图片，但当前模型不支持图片理解。请友好地提示用户切换到支持视觉能力的模型（如 qwen-vl-plus / qwen-vl-max）后重试。"})

    # 文档文本注入
    if doc_text_parts and not (image_attachments and has_vision):
        print(f"[Chat] 注入文档上下文: {len(doc_text_parts)} 个文件")
        ai_messages.insert(-1, {"role": "system", "content": "用户附带了以下文件内容，请结合文件内容回答用户的问题：\n\n" + "\n\n".join(doc_text_parts)})

    # --- SSE 流式生成器 ---
    import json as _json
    search_res_json = None  # 文件搜索结果（由循环内工具调用填充）

    def _sse_generator():
        from datetime import datetime as _dt
        import time as _time

        thinking_buf = ""
        reply_buf = ""

        mcp_intent = intent.get("mcp_intent", "NONE")
        mcp_tools = mcp_mgr.get_all_tools_for_loop(mcp_intent)
        current_messages = list(ai_messages)

        # 注入当前时间（让 AI 知道现在几点）
        _now_str = _dt.now().strftime("%Y年%m月%d日 %H:%M（%A）")
        current_messages.insert(1, {
            "role": "system",
            "content": f"当前时间：{_now_str}",
        })

        # ====== Plan-Act-Observe 循环配置 ======
        MAX_LOOPS = int(APP_SETTINGS.get("max_tool_loops", 6))
        turn_number = session_mgr.get_turn_number()
        tool_summary_parts: list[str] = []  # 收集工具调用摘要（写入 chat.md）
        render_cards: list[dict] = []       # 收集渲染用工具卡片
        render_media: list[str] = []        # 收集渲染用媒体路径

        # 记录用户消息到 ai_context.jsonl
        session_mgr.append_ai_context({
            "turn": turn_number,
            "role": "user",
            "content": user_text[:500],
            "ts": _dt.now().isoformat(),
        })

        # ====== 构建内置工具列表（RAG 搜索 + 文件搜索） ======
        builtin_tools = [
            {
                "type": "function",
                "function": {
                    "name": "_builtin_rag_search",
                    "description": "搜索用户私有的本地知识库（仅包含用户主动上传的笔记、课件、技术文档等参考资料）。注意：这不是互联网搜索，仅当用户明确要求查询自己上传的资料时才调用。对于通用问题（如天气、新闻、美食推荐等）请使用联网搜索工具。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "query": {"type": "string", "description": "搜索关键词，3-5 个核心词"}
                        },
                        "required": ["query"]
                    }
                }
            },
        ]
        if minio_mgr.enabled:
            builtin_tools.append({
                "type": "function",
                "function": {
                    "name": "_builtin_file_search",
                    "description": "在云端网盘中搜索用户的文件、照片、文档、音乐等。当用户想查找或定位文件时调用。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "query": {"type": "string", "description": "文件查找描述，如'去年的旅行照片'、'工作周报'"}
                        },
                        "required": ["query"]
                    }
                }
            })
            builtin_tools.append({
                "type": "function",
                "function": {
                    "name": "_builtin_file_upload",
                    "description": "将用户在聊天中发送的附件文件保存到云端网盘。当用户明确表示要存储/保存/上传附件到网盘时调用。需要用户在消息中附带文件。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "description": {"type": "string", "description": "文件描述，用于后续搜索和标签生成，从用户消息中提取关键信息"}
                        },
                        "required": ["description"]
                    }
                }
            })

        # 合并所有工具（内置 + MCP）
        all_tools = builtin_tools + (mcp_tools if mcp_tools else [])

        if mcp_tools:
            current_messages.insert(0, {
                "role": "system",
                "content": "【重要指令】你已连接外部 MCP 工具。当用户要求画图、搜索或执行任务时，你**必须严格优先调用对应的 Tool**（如 modelstudio_z_image_generation），**绝对禁止**凭空伪造图片链接！只有在你确切执行了工具后即可，后续展示由前端强制完成，你只需文字总结即可。"
            })

        # 注入工具使用指引
        current_messages.insert(0, {
            "role": "system",
            "content": (
                "你拥有以下内置能力，请在需要时主动调用：\n"
                "1. _builtin_rag_search: 搜索本地知识库（用户上传的笔记/资料）\n"
                + ("2. _builtin_file_search: 在云端网盘中查找文件/照片/文档\n"
                   "3. _builtin_file_upload: 将聊天附件保存到云端网盘（用户需附带文件）\n" if minio_mgr.enabled else "")
                + "\n【Plan→Act→Observe 工作流】\n"
                "每次回复时，你必须遵循以下流程：\n"
                "1. **Plan（规划）**: 先用 1-2 句话简要说明你的分析和下一步计划，这段文字会展示给用户看\n"
                "2. **Act（执行）**: 调用工具获取信息或生成内容\n"
                "3. **Observe（观察）**: 工具返回结果后，分析其内容，决定是否需要更多操作\n"
                "\n重要：每次调用工具之前，你必须先输出“💡 分析”前缀的简短思考。例如：\n"
                "“💡 分析：用户要搜索南京美食然后画图，我先用联网搜索找具体美食名称…”\n"
                "\n如果用户问题不需要查资料或找文件，直接回答即可，不必调用工具。\n"
                f"你最多可以进行 {MAX_LOOPS} 轮工具调用循环，请合理规划。"
            )
        })

        executed_tools_ui = set()  # Track which tool UI cards have been emitted
        nonlocal search_res_json  # 允许循环内填充文件搜索结果

        for loop_round in range(1, MAX_LOOPS + 1):
            remaining = MAX_LOOPS - loop_round + 1

            # 推送循环进度状态给前端
            if loop_round > 1:
                yield f"event: loop_status\ndata: {_json.dumps({'loop': loop_round, 'remaining': remaining, 'max': MAX_LOOPS}, ensure_ascii=False)}\n\n"

            # 注入剩余步数提示（仅在已执行过工具后的后续轮次注入）
            if loop_round > 1:
                step_hint = {
                    "role": "system",
                    "content": (
                        f"\n🔍 Observe — 第 {loop_round} 轮（剩余 {remaining} 步）\n"
                        f"上方的 tool 消息是上一轮工具的返回结果。请你：\n"
                        f"1. 先输出一段“💡 分析：...”简要说明你从结果中发现了什么\n"
                        f"2. 然后决定下一步动作：调用更多工具 或 用获得的信息直接回复用户\n"
                        f"\n示例：\n"
                        f'“💡 分析：搜索结果显示南京最高的山是紫金山（海拔448.9m），现在用这个具体信息生成图片…”'
                    ),
                }
                current_messages.append(step_hint)

            # 记录 Plan 到 ai_context
            session_mgr.append_ai_context({
                "turn": turn_number,
                "role": "plan",
                "loop": loop_round,
                "remaining": remaining,
                "has_tools": bool(all_tools),
                "ts": _dt.now().isoformat(),
            })

            skip_temp = "reasoning" in model_caps or "fixed_temp" in model_caps
            stream_kwargs = {
                "model": current_model,
                "messages": current_messages,
                "stream": True
            }
            if not skip_temp:
                stream_kwargs["temperature"] = 0.7
            if all_tools:
                stream_kwargs["tools"] = all_tools

            try:
                stream = client.chat.completions.create(**stream_kwargs)
            except Exception as e:
                # 某些模型不支持 tools 参数（如 MiniMax），降级为无工具模式重试
                if "400" in str(e) and "tools" in stream_kwargs:
                    print(f"[Chat] 模型不支持 tools 参数，降级为无工具模式: {e}")
                    stream_kwargs.pop("tools", None)
                    all_tools = []  # 本轮及后续轮次不再注入工具
                    try:
                        stream = client.chat.completions.create(**stream_kwargs)
                    except Exception as e2:
                        session_mgr.pop_last_message()
                        error_type = type(e2).__name__
                        yield f"event: error\ndata: {_json.dumps({'message': f'{error_type}: {str(e2)}'}, ensure_ascii=False)}\n\n"
                        return
                else:
                    session_mgr.pop_last_message()
                    error_type = type(e).__name__
                    yield f"event: error\ndata: {_json.dumps({'message': f'{error_type}: {str(e)}'}, ensure_ascii=False)}\n\n"
                    return

            tool_calls_buffer = {}
            has_tool_call = False

            for chunk in stream:
                delta = chunk.choices[0].delta if chunk.choices else None
                if not delta:
                    continue

                # 1. 缓冲 tool_calls
                if getattr(delta, "tool_calls", None):
                    has_tool_call = True
                    for tc in delta.tool_calls:
                        idx = tc.index
                        if idx not in tool_calls_buffer:
                            tool_calls_buffer[idx] = {"id": tc.id, "type": "function", "function": {"name": tc.function.name, "arguments": ""}}
                        if getattr(tc.function, "arguments", None):
                            # type: ignore
                            tool_calls_buffer[idx]["function"]["arguments"] += tc.function.arguments

                # 检测 reasoning_content（推理思考内容，兼容多家模型）
                reasoning = (
                    getattr(delta, "reasoning_content", None)
                    or getattr(delta, "thinking", None)
                    or getattr(delta, "thinking_content", None)
                    or ""
                )
                if reasoning:
                    thinking_buf += reasoning  # type: ignore[operator]
                    yield f"event: thinking\ndata: {_json.dumps({'text': reasoning}, ensure_ascii=False)}\n\n"

                # 正文内容
                content = delta.content or ""
                if content:
                    reply_buf += content  # type: ignore[operator]
                    yield f"event: delta\ndata: {_json.dumps({'text': content}, ensure_ascii=False)}\n\n"
            
            # --- 单轮流结束，检查是否执行了工具 ---
            if has_tool_call and tool_calls_buffer:
                # 记录 assistant 消息（包含 Plan 分析文本 + tool_calls）
                # 保留 AI 的推理文本，使其在后续轮次中可见
                plan_text = reply_buf.strip() if reply_buf.strip() else None

                # 从 reply_buf 中删除中间推理文本（Plan/分析），只保留工具卡片和媒体 HTML
                import re as _re_strip
                _tools_html = ""
                _tools_match = _re_strip.search(r'<div class="tool-cards-row">.*?</div><!-- END_TOOLS -->', reply_buf, flags=_re_strip.DOTALL)
                if _tools_match:
                    _tools_html = _tools_match.group(0)
                _media_html = ""
                _media_match = _re_strip.search(r'<div class="media-grid">.*?</div><!-- END_MEDIA_GRID -->', reply_buf, flags=_re_strip.DOTALL)
                if _media_match:
                    _media_html = _media_match.group(0)
                _video_html = ""
                _video_match = _re_strip.search(r'<div class="media-video">.*?</div>', reply_buf, flags=_re_strip.DOTALL)
                if _video_match:
                    _video_html = _video_match.group(0)
                reply_buf = ""
                if _tools_html:
                    reply_buf += f"\n{_tools_html}\n"
                if _media_html:
                    reply_buf += f"\n{_media_html}\n"
                if _video_html:
                    reply_buf += f"\n{_video_html}\n"

                assistant_tc_msg: dict = {
                    "role": "assistant",
                    "content": plan_text,
                    "tool_calls": list(tool_calls_buffer.values())
                }
                current_messages.append(assistant_tc_msg)

                # 逐个执行工具
                for tc in tool_calls_buffer.values():
                    func_dict2 = tc["function"]
                    t_name = func_dict2["name"]  # type: ignore[index]
                    t_args_str = func_dict2["arguments"]  # type: ignore[index]
                    try:
                        args_dict = _json.loads(t_args_str)
                    except:
                        args_dict = {}
                    
                    friendly_name = {
                        "_builtin_rag_search": "知识库检索",
                        "_builtin_file_search": "网盘文件搜索",
                        "modelstudio_z_image_generation": "Z-Image 图像生成",
                        "amap_poi_search": "高德地图路线",
                        "zhipu_web_search": "智谱联网搜索",
                        "qwen_tts": "智能语音合成",
                        "web_search": "Bing 联网搜索",
                        "webSearchPro": "联网搜索 Pro",
                        "webSearchStd": "联网搜索",
                        "webSearchSogou": "搜狗搜索",
                        "webSearchQuark": "夸克搜索",
                        "bailian_web_search": "百炼联网搜索",
                        "jimeng_image_generation": "即梦 AI 图片生成",
                        "jimeng_video_generation": "即梦 AI 视频生成",
                    }.get(t_name, t_name.replace("_", " ").title())
                    
                    working_card = f'<div class="tool-call-card working"><div class="tool-spinner"></div>正在调用技能：{friendly_name}</div>'
                    success_card = f'<div class="tool-call-card success">✅ 成功执行技能：{friendly_name}</div>'

                    TOOLS_ROW_START = '<div class="tool-cards-row">'
                    TOOLS_ROW_END = '</div><!-- END_TOOLS -->'
                    
                    # Revert success to working if it's already generated previously
                    if friendly_name not in executed_tools_ui:
                        # 首次出现此工具 — 添加到容器
                        if TOOLS_ROW_START not in reply_buf:
                            # 创建容器 + 放入第一个卡片
                            reply_buf = f'\n{TOOLS_ROW_START}{working_card}{TOOLS_ROW_END}\n' + reply_buf
                        else:
                            # 追加到容器内
                            reply_buf = reply_buf.replace(TOOLS_ROW_END, f'{working_card}{TOOLS_ROW_END}')
                        executed_tools_ui.add(friendly_name)
                        yield f"event: replace_all\ndata: {_json.dumps({'text': reply_buf}, ensure_ascii=False)}\n\n"
                    else:
                        # Revert the success card to working card temporarily
                        if success_card in reply_buf:
                            reply_buf = reply_buf.replace(success_card, working_card)
                            yield f"event: replace_all\ndata: {_json.dumps({'text': reply_buf}, ensure_ascii=False)}\n\n"
                    
                    t_start = _time.time()
                    # 分流执行：内置工具 vs MCP 工具
                    if t_name == "_builtin_rag_search":
                        query = args_dict.get("query", "")
                        rag_result = rag.retrieve_context(query, top_k=3)
                        if rag_result:
                            t_result = "知识库检索结果：\n" + rag_result
                        else:
                            t_result = "未在知识库中找到相关内容。"
                    elif t_name == "_builtin_file_search":
                        query = args_dict.get("query", "")
                        _search_res = minio_mgr.ai_search(query)
                        if _search_res.get("status") == "ok" and _search_res.get("files"):
                            search_res_json = _search_res  # 填充到外部变量
                            t_result = (
                                f"已找到匹配文件，将以卡片形式附加在回复下方。"
                                f"匹配原因：{_search_res.get('reason', '')}。"
                                f"请仅用一两句话告知用户找到了文件，**不要**展示文件名或下载链接。"
                            )
                        else:
                            t_result = f"在网盘中未找到描述为\u201c{query}\u201d的文件。请告知用户未找到。"
                    elif t_name == "_builtin_file_upload":
                        desc = args_dict.get("description", "")
                        if not attachments:
                            t_result = "用户没有在消息中附带任何文件，无法上传。请提醒用户在发送消息时附加文件。"
                        else:
                            import base64 as _b64_upload
                            uploaded_names = []
                            for att in attachments:
                                att_name = att.get("name", "unnamed")
                                att_mime = att.get("mime", "application/octet-stream")
                                att_data = _b64_upload.b64decode(att.get("data", ""))
                                if not att_data:
                                    continue
                                try:
                                    entry = minio_mgr.upload_fast(
                                        filename=att_name,
                                        file_data=att_data,
                                        content_type=att_mime,
                                        description=desc,
                                    )
                                    uploaded_names.append(f"{att_name} → {entry['object_name']}")
                                    # 后台处理标签
                                    _obj = entry["object_name"]
                                    _ct = att_mime
                                    def _bg_tag(_o=_obj, _n=att_name, _c=_ct, _d=desc):
                                        try:
                                            result = minio_mgr.process_tags(_o, _n, _c, _d)
                                            broadcast_sync({
                                                "type": "file_tags_ready",
                                                "object_name": _o,
                                                "original_name": _n,
                                                "tags": result["tags"],
                                                "categorized_tags": result.get("categorized_tags", {}),
                                                "file_meta": result.get("file_meta", {}),
                                            })
                                        except Exception as _e:
                                            print(f"[ChatUpload] 标签处理失败: {_e}")
                                    threading.Thread(target=_bg_tag, daemon=True).start()
                                except Exception as ue:
                                    uploaded_names.append(f"{att_name} → 上传失败: {ue}")
                            t_result = f"已将 {len(uploaded_names)} 个文件保存到云端网盘：{'; '.join(uploaded_names)}。标签正在后台生成中。请告知用户文件已保存成功。"
                    else:
                        # 根据工具名自动路由到正确的 MCP intent
                        _TOOL_TO_INTENT = {
                            "web_search": "WEB_SEARCH",
                            "webSearchPro": "WEB_SEARCH",
                            "webSearchStd": "WEB_SEARCH",
                            "webSearchSogou": "WEB_SEARCH",
                            "webSearchQuark": "WEB_SEARCH",
                            "jimeng_image_generation": "JIMENG",
                            "jimeng_video_generation": "JIMENG",
                            "modelstudio_z_image_generation": "Z_IMAGE",
                            "amap_poi_search": "AMAP",
                            "maps_weather": "AMAP",
                            "qwen_tts": "TTS",
                        }
                        tool_intent = _TOOL_TO_INTENT.get(t_name, mcp_intent)
                        t_result = mcp_mgr.execute_tool(
                            intent=tool_intent,
                            tool_name=t_name,
                            args=args_dict,
                            session_id=session_mgr.session_id,
                            session_dir=session_mgr.session_dir
                        )
                    t_duration_ms = int((_time.time() - t_start) * 1000)
                    
                    t_result_str = str(t_result)

                    # 记录工具观察到 ai_context.jsonl（只存有用信息，不存调用机制）
                    # 提取简短信息摘要
                    import re as _re_ctx
                    _info = t_result_str[:300]
                    # 去掉 HTML 标签，只保留文本信息
                    _info = _re_ctx.sub(r'<[^>]+>', '', _info)
                    # 去掉【重要指令】等内部钩子
                    _info = _re_ctx.sub(r'【重要指令】.*', '', _info)
                    _info = _info.strip()
                    session_mgr.append_ai_context({
                        "turn": turn_number,
                        "role": "observation",
                        "loop": loop_round,
                        "source": friendly_name,
                        "info": _info if _info else "ⓘ 无文本返回（媒体已生成）",
                        "duration_ms": t_duration_ms,
                        "ts": _dt.now().isoformat(),
                    })

                    # 收集工具摘要和渲染数据
                    tool_summary_parts.append(f"{friendly_name} → ✅")
                    render_cards.append({"name": friendly_name, "status": "success"})
                    
                    # Intercept the HTML from mcp_manager and force layout it without LLM involvement
                    import re
                    img_match = re.search(r'<a href="/assets/[^>]+><img[^>]+></a>', t_result_str)
                    vid_match = re.search(r'<video[^>]+src="[^"]+/assets/[^"]+"[^>]*></video>', t_result_str)

                    if img_match:
                        img_html = img_match.group(0)
                        grid_start = '<div class="media-grid">'
                        grid_end_marker = '</div><!-- END_MEDIA_GRID -->'
                        
                        if grid_start not in reply_buf:
                            reply_buf += f"\n{grid_start}\n{img_html}\n{grid_end_marker}\n"
                        else:
                            reply_buf = reply_buf.replace(grid_end_marker, f"{img_html}\n{grid_end_marker}")

                        # 收集媒体路径用于 render.jsonl
                        _img_path_m = re.search(r'href="(/assets/[^"]+)"', img_html)
                        if _img_path_m:
                            render_media.append(_img_path_m.group(1))

                        # Strip the raw instruction hook out of the LLM context so it doesn't try to regurgitate it
                        context_msg = "✅ 工具调用成功！图片已在界面直接展示给用户，请简要用一两句话评价生成的图片即可，绝对不要再回复任何图片链接。"
                    elif vid_match:
                        vid_html = vid_match.group(0)
                        reply_buf += f"\n<div class=\"media-video\">\n{vid_html}\n</div>\n"
                        # 收集视频路径
                        _vid_path_m = re.search(r'src="(/assets/[^"]+)"', vid_html)
                        if _vid_path_m:
                            render_media.append(_vid_path_m.group(1))
                        context_msg = "✅ 工具调用成功！视频已在界面直接展示给用户，请简要用一两句话描述生成的视频内容即可，绝对不要再回复任何视频链接。"
                    else:
                        context_msg = t_result_str
                        
                    if working_card in reply_buf:
                        reply_buf = reply_buf.replace(working_card, success_card)
                        
                    yield f"event: replace_all\ndata: {_json.dumps({'text': reply_buf}, ensure_ascii=False)}\n\n"
                    
                    current_messages.append({
                        "role": "tool",
                        "tool_call_id": tc["id"],
                        "name": t_name,
                        "content": context_msg
                    })
                
                print(f"[Chat] Plan-Act-Observe 循环第 {loop_round} 轮完成，剩余 {remaining - 1} 步")
                # 携带工具结果进行下一轮请求
                continue
            else:
                # 无工具调用，流完全结束
                break

        # --- 流结束后的后处理 ---
        # 构建工具调用摘要字符串
        tool_summary = " | ".join(tool_summary_parts) if tool_summary_parts else ""
        session_mgr.append_ai_message(reply_buf, thinking=thinking_buf, tool_summary=tool_summary)

        # 写入渲染事件到 render.jsonl
        if render_cards:
            session_mgr.append_render_event({
                "turn": turn_number, "type": "tool_cards", "cards": render_cards
            })
        if render_media:
            session_mgr.append_render_event({
                "turn": turn_number, "type": "media", "paths": render_media
            })
        if search_res_json and search_res_json.get("files"):
            session_mgr.append_render_event({
                "turn": turn_number, "type": "file_search", "data": search_res_json
            })

        # 记录 AI 最终回复到 ai_context.jsonl（纯文本，不含 HTML）
        clean_reply = SessionManager._strip_html_for_chat(reply_buf)
        session_mgr.append_ai_context({
            "turn": turn_number,
            "role": "assistant",
            "content": clean_reply[:500],
            "thinking": thinking_buf[:200] if thinking_buf else "",
            "tool_count": len(tool_summary_parts),
            "ts": _dt.now().isoformat(),
        })

        if session_mgr.is_first_message:
            session_mgr.mark_first_message_done()
        else:
            session_mgr.update_activity()

        # 构建 done 事件的附加数据
        done_data: dict = {}

        # 自动任务
        if task_intent:
            try:
                created = task_scheduler.create_task(
                    task_name=task_intent["task_name"],
                    trigger_type=task_intent["trigger_type"],
                    trigger_args=task_intent["trigger_args"],
                    action_prompt=task_intent["action_prompt"],
                )
                broadcast_sync({"type": "task_created", "task": created})
                session_mgr.append_event("auto_task", created)
                done_data["auto_task"] = created
            except Exception as e:
                print(f"[AutoTask] 自动创建任务失败: {e}")

        # 自动创建日程
        if schedule_intent and schedule_intent.get("action") == "create":
            try:
                sch_data = {
                    "title": schedule_intent["title"],
                    "start_time": schedule_intent["start_time"],
                    "end_time": schedule_intent.get("end_time", ""),
                    "description": schedule_intent.get("description", ""),
                    "category": schedule_intent.get("category", "其他"),
                    "location": schedule_intent.get("location", ""),
                    "all_day": schedule_intent.get("all_day", False),
                }
                # 如果没有结束时间，默认 1 小时
                if not sch_data["end_time"]:
                    from datetime import datetime as _dt, timedelta as _td
                    start = _dt.fromisoformat(sch_data["start_time"])
                    sch_data["end_time"] = (start + _td(hours=1)).isoformat()
                created_sch = schedule_mgr.create(sch_data)
                broadcast_sync({"type": "schedule_created", "schedule": created_sch})
                session_mgr.append_event("auto_schedule", created_sch)
                done_data["auto_schedule"] = created_sch
            except Exception as e:
                print(f"[AutoSchedule] 自动创建日程失败: {e}")

        # 文件搜索结果
        if search_res_json:
            session_mgr.append_event("file_search", search_res_json)
            done_data["file_search_result"] = search_res_json

        # 持久化附件
        if attachments:
            import base64 as _b64
            assets_dir = os.path.join(session_mgr.session_dir, "assets")
            os.makedirs(assets_dir, exist_ok=True)
            saved_files: list[dict] = []
            for att in attachments:
                ext = os.path.splitext(att.get("name", ""))[1] or ".bin"
                unique_name = _uuid.uuid4().hex[:8] + ext  # type: ignore[index]
                filepath = os.path.join(assets_dir, unique_name)
                try:
                    raw = _b64.b64decode(att.get("data", ""))
                    with open(filepath, "wb") as f:
                        f.write(raw)
                    saved_files.append({
                        "filename": unique_name,
                        "original_name": att.get("name", "未知"),
                        "mime": att.get("mime", "application/octet-stream"),
                        "url": f"/api/session_asset/{session_mgr.session_id}/{unique_name}",
                    })
                except Exception as e:
                    print(f"[Chat] 保存附件失败: {e}")
            if saved_files:
                session_mgr.append_event("chat_attach", {"files": saved_files}, after_msg_index=user_msg_index)
                done_data["chat_attachments"] = saved_files

        yield f"event: done\ndata: {_json.dumps(done_data, ensure_ascii=False)}\n\n"

    return StreamingResponse(_sse_generator(), media_type="text/event-stream")




# ====== 会话资源路由 ======

@app.get("/api/session_asset/{session_id}/{filename}")
def serve_session_asset(session_id: str, filename: str):
    """提供会话中保存的附件文件（图片、PDF 等）"""
    from fastapi import HTTPException  # type: ignore[import]
    filepath = os.path.join(SESSION_DIR, session_id, "assets", filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Asset not found")
    return FileResponse(filepath)

# ====== 记忆与系统提示词管理路由 ======

@app.get("/api/memory")
def get_memory():
    """获取 memory.md 内容"""
    content = ""
    if os.path.exists(MEMORY_FILE):
        try:
            with open(MEMORY_FILE, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            return {"status": "error", "message": str(e)}
    return {"status": "ok", "content": content}

class MemoryUpdateRequest(BaseModel):
    content: str

@app.post("/api/memory")
def update_memory(req: MemoryUpdateRequest):
    """保存 memory.md 内容"""
    try:
        os.makedirs(os.path.dirname(MEMORY_FILE), exist_ok=True)
        with open(MEMORY_FILE, "w", encoding="utf-8") as f:
            f.write(req.content)
        return {"status": "ok", "message": "记忆已保存"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/system_prompt")
def get_system_prompt():
    """获取 system_prompt.md 内容"""
    from backend.config import SYSTEM_PROMPT_FILE  # type: ignore[import]
    content = ""
    if os.path.exists(SYSTEM_PROMPT_FILE):
        try:
            with open(SYSTEM_PROMPT_FILE, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            return {"status": "error", "message": str(e)}
    return {"status": "ok", "content": content}

class SystemPromptUpdateRequest(BaseModel):
    content: str

@app.post("/api/system_prompt")
def update_system_prompt(req: SystemPromptUpdateRequest):
    """保存 system_prompt.md 并更新运行时提示词"""
    from backend.config import SYSTEM_PROMPT_FILE  # type: ignore[import]
    import backend.config as _cfg  # type: ignore[import]
    try:
        os.makedirs(os.path.dirname(SYSTEM_PROMPT_FILE), exist_ok=True)
        with open(SYSTEM_PROMPT_FILE, "w", encoding="utf-8") as f:
            f.write(req.content)
        # 更新运行时的 SYSTEM_PROMPT（后续新会话会使用新提示词）
        _cfg.SYSTEM_PROMPT = req.content
        return {"status": "ok", "message": "系统提示词已保存，新会话将使用更新后的提示词"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ====== 定时任务路由 ======


class TaskCreateRequest(BaseModel):
    task_name: str
    trigger_type: str           # "date" | "interval" | "cron"
    trigger_args: dict
    action_prompt: str

@app.get("/api/tasks")
def get_tasks():
    return {"tasks": task_scheduler.list_tasks()}

@app.post("/api/tasks")
def create_task(req: TaskCreateRequest):
    task = task_scheduler.create_task(
        task_name=req.task_name,
        trigger_type=req.trigger_type,
        trigger_args=req.trigger_args,
        action_prompt=req.action_prompt,
    )
    return {"status": "ok", "task": task}

@app.post("/api/tasks/{task_id}/pause")
def pause_task(task_id: str):
    ok = task_scheduler.pause_task(task_id)
    return {"status": "ok" if ok else "error"}

@app.post("/api/tasks/{task_id}/resume")
def resume_task(task_id: str):
    ok = task_scheduler.resume_task(task_id)
    return {"status": "ok" if ok else "error"}

@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: str):
    ok = task_scheduler.delete_task(task_id)
    return {"status": "ok" if ok else "error"}

# ====== 日程管理路由 ======

class ScheduleCreateRequest(BaseModel):
    title: str
    start_time: str
    end_time: str
    description: str = ""
    all_day: bool = False
    category: str = "其他"
    color: str = ""
    location: str = ""
    rrule: str = ""
    reminder_minutes: int = 15

class ScheduleUpdateRequest(BaseModel):
    title: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    description: Optional[str] = None
    all_day: Optional[bool] = None
    category: Optional[str] = None
    color: Optional[str] = None
    location: Optional[str] = None
    rrule: Optional[str] = None
    reminder_minutes: Optional[int] = None
    status: Optional[str] = None

class ScheduleParseRequest(BaseModel):
    text: str

@app.get("/api/schedules")
def list_schedules(start: str = Query(""), end: str = Query("")):
    """按时间范围查询日程"""
    if not start or not end:
        return {"schedules": schedule_mgr.list_all()}
    return {"schedules": schedule_mgr.list_range(start, end)}

@app.post("/api/schedules")
def create_schedule(req: ScheduleCreateRequest):
    """创建日程"""
    schedule = schedule_mgr.create(req.model_dump())
    return {"status": "ok", "schedule": schedule}

@app.put("/api/schedules/{schedule_id}")
def update_schedule(schedule_id: str, req: ScheduleUpdateRequest):
    """更新日程"""
    data = {k: v for k, v in req.model_dump().items() if v is not None}
    result = schedule_mgr.update(schedule_id, data)
    if result is None:
        return {"status": "error", "message": "日程不存在"}
    return {"status": "ok", "schedule": result}

@app.delete("/api/schedules/{schedule_id}")
def delete_schedule(schedule_id: str):
    """删除日程"""
    ok = schedule_mgr.delete(schedule_id)
    return {"status": "ok" if ok else "error"}

@app.get("/api/schedules/today")
def get_today_briefing():
    """获取今日日程摘要"""
    briefing = schedule_mgr.generate_daily_briefing()
    from datetime import date as _date
    today = _date.today().isoformat()
    schedules = schedule_mgr.list_range(today, today)
    return {"briefing": briefing, "schedules": schedules}

@app.post("/api/schedules/parse")
def parse_schedule_text(req: ScheduleParseRequest):
    """AI 解析自然语言为日程 JSON"""
    result = schedule_mgr.parse_natural_language(req.text)
    if result:
        return {"status": "ok", "schedule_data": result}
    return {"status": "error", "message": "无法解析日程信息"}

# ====== 通知通道路由 ======

@app.get("/api/notifications")
def get_notification_config():
    """获取当前通知通道配置"""
    config = load_notification_config()
    # 脱敏 app_secret
    channels = config.get("channels", {})
    dt = channels.get("dingtalk", {})
    secret = dt.get("app_secret", "")
    if secret and len(secret) > 10:
        dt["app_secret_masked"] = f"{secret[:6]}...{secret[-4:]}"
    else:
        dt["app_secret_masked"] = "未配置"
    dt.pop("app_secret", None)
    return {"config": config}

class NotificationConfigRequest(BaseModel):
    channels: dict

@app.post("/api/notifications")
def update_notification_config(req: NotificationConfigRequest):
    """更新通知通道配置（重启后生效）"""
    current = load_notification_config()
    for ch_name, ch_cfg in req.channels.items():
        if ch_name in current["channels"]:
            current["channels"][ch_name].update(ch_cfg)
        else:
            current["channels"][ch_name] = ch_cfg
    save_notification_config(current)
    return {"status": "ok", "message": "配置已保存，重启服务后生效"}

@app.post("/api/notifications/test")
def test_notification():
    """发送测试消息到钉钉，验证通道连通性"""
    if notification_mgr is None:
        return {"status": "error", "message": "通知管理器未初始化"}
    dt_channel = notification_mgr.get_channel("dingtalk")
    if dt_channel is None or not dt_channel.enabled:
        return {"status": "error", "message": "钉钉通道未启用"}
    # DingTalkChannel 类型断言
    from backend.notification_manager import DingTalkChannel as _DT  # type: ignore[import]
    if isinstance(dt_channel, _DT):
        ok = dt_channel.test_send()
        if ok:
            return {"status": "ok", "message": "测试消息已发送，请检查钉钉"}
        return {"status": "error", "message": "发送失败，请检查配置和日志"}
    return {"status": "error", "message": "通道类型错误"}

# ====== 文件管理路由 ======

# 上传临时目录（流式接收大文件用）
UPLOAD_TEMP_DIR = os.path.join("data", "upload_temp")
os.makedirs(UPLOAD_TEMP_DIR, exist_ok=True)

@app.post("/api/files/upload")
async def upload_file(
    file: UploadFile = File(...),
    description: str = Form(""),
):
    """上传文件到 MinIO（流式写入临时文件，不占用大量内存）"""
    if not minio_mgr.enabled:
        return {"status": "error", "message": "MinIO 未配置"}

    import tempfile
    filename = file.filename or "unnamed"
    content_type = file.content_type or "application/octet-stream"
    ext = os.path.splitext(filename)[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext, dir=UPLOAD_TEMP_DIR)

    try:
        # 1. 流式写入临时文件（每次只占 1MB 内存）
        file_size = 0
        while chunk := await file.read(1024 * 1024):  # 1MB chunks
            tmp.write(chunk)
            file_size += len(chunk)
        tmp.close()

        # 2. 从本地文件直接上传到 MinIO（零内存拷贝）
        entry = minio_mgr.upload_fast_from_file(
            filename=filename,
            file_path=tmp.name,
            file_size=file_size,
            content_type=content_type,
            description=description,
        )

        # 3. 后台线程处理元信息提取 + AI 打标签
        object_name = entry["object_name"]

        def _background_tagging():
            try:
                result = minio_mgr.process_tags(
                    object_name=object_name,
                    filename=filename,
                    content_type=content_type,
                    description=description,
                )
                broadcast_sync({
                    "type": "file_tags_ready",
                    "object_name": object_name,
                    "original_name": filename,
                    "tags": result["tags"],
                    "categorized_tags": result.get("categorized_tags", {}),
                    "file_meta": result.get("file_meta", {}),
                })
            except Exception as e:
                print(f"[AsyncTag] 后台标签处理失败: {e}")
                broadcast_sync({
                    "type": "file_tags_ready",
                    "object_name": object_name,
                    "original_name": filename,
                    "tags": [],
                    "error": str(e),
                })

        threading.Thread(target=_background_tagging, daemon=True).start()

        return {"status": "ok", "file": entry}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        # 清理临时文件
        try:
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
        except Exception:
            pass


@app.post("/api/files/upload_batch")
async def upload_files_batch(
    files: list[UploadFile] = File(...),
    description: str = Form(""),
):
    """批量上传多个文件到 MinIO（流式写入临时文件，不占用大量内存）"""
    if not minio_mgr.enabled:
        return {"status": "error", "message": "MinIO 未配置"}
    if not files:
        return {"status": "error", "message": "未选择任何文件"}

    import tempfile
    tmp_paths: list[str] = []  # 用于 finally 清理

    try:
        # 1. 流式写入临时文件（每个文件每次只占 1MB 内存）
        file_tuples: list[tuple[str, str, int, str, str]] = []
        for f in files:
            fname = f.filename or "unnamed"
            ctype = f.content_type or "application/octet-stream"
            ext = os.path.splitext(fname)[1]
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext, dir=UPLOAD_TEMP_DIR)
            tmp_paths.append(tmp.name)

            file_size = 0
            while chunk := await f.read(1024 * 1024):  # 1MB chunks
                tmp.write(chunk)
                file_size += len(chunk)
            tmp.close()

            file_tuples.append((fname, tmp.name, file_size, ctype, description))

        # 2. 批量快速上传（从本地文件直接上传到 MinIO）
        entries = minio_mgr.upload_batch_from_files(file_tuples)

        # 3. 后台线程逐个处理元信息 + AI 标签
        def _background_batch_tagging():
            for i, entry in enumerate(entries):
                try:
                    result = minio_mgr.process_tags(
                        object_name=entry["object_name"],
                        filename=entry["original_name"],
                        content_type=entry["content_type"],
                        description=entry.get("description", ""),
                    )
                    broadcast_sync({
                        "type": "file_tags_ready",
                        "object_name": entry["object_name"],
                        "original_name": entry["original_name"],
                        "tags": result["tags"],
                        "categorized_tags": result.get("categorized_tags", {}),
                        "file_meta": result.get("file_meta", {}),
                        "batch_progress": f"{i + 1}/{len(entries)}",
                    })
                except Exception as e:
                    print(f"[AsyncTag] 批量标签处理失败 ({entry['original_name']}): {e}")
                    broadcast_sync({
                        "type": "file_tags_ready",
                        "object_name": entry["object_name"],
                        "original_name": entry["original_name"],
                        "tags": [],
                        "error": str(e),
                        "batch_progress": f"{i + 1}/{len(entries)}",
                    })

        threading.Thread(target=_background_batch_tagging, daemon=True).start()

        return {
            "status": "ok",
            "total": len(file_tuples),
            "uploaded": len(entries),
            "failed": len(file_tuples) - len(entries),
            "files": entries,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        # 清理所有临时文件
        for p in tmp_paths:
            try:
                if os.path.exists(p):
                    os.unlink(p)
            except Exception:
                pass

@app.get("/api/files/search")
def search_files(q: str = Query("")):
    """搜索文件（按文件名/描述/标签）"""
    results = minio_mgr.search(q)
    # 为每个结果生成下载链接，强制开启下载模式并指定原始文件名
    for r in results:
        r["download_url"] = minio_mgr.get_download_url(
            r["object_name"], force_download=True, filename=r.get("original_name")
        )
    return {"files": results}

class AISearchRequest(BaseModel):
    prompt: str

@app.post("/api/files/ai_search")
def ai_search_files(req: AISearchRequest):
    """AI 语义检索文件（基于标签池匹配）"""
    if not minio_mgr.enabled:
        return {"status": "error", "message": "MinIO 未配置"}
    try:
        result = minio_mgr.ai_search(req.prompt)
        return result
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/files/tags")
def get_file_tags():
    """获取全局标签池（按分类汇总）"""
    if not minio_mgr.enabled:
        return {"status": "error", "message": "MinIO 未配置"}
    return {"status": "ok", "tags": minio_mgr.get_tag_pool()}

@app.get("/api/files")
def list_files():
    """列出所有文件"""
    files = minio_mgr.list_files()
    for f in files:
        f["download_url"] = minio_mgr.get_download_url(
            f["object_name"], force_download=True, filename=f.get("original_name")
        )
    return {"files": files}

@app.delete("/api/files/{object_name:path}")
def delete_file(object_name: str):
    """删除文件"""
    ok = minio_mgr.delete(object_name)
    if ok:
        return {"status": "ok"}
    return {"status": "error", "message": "删除失败"}

# ====== 音乐功能路由 ======

class PlaylistRequest(BaseModel):
    prompt: str

@app.get("/api/music")
def list_music():
    """列出所有音频文件"""
    try:
        audio_files = minio_mgr.search_audio()
        for f in audio_files:
            f["download_url"] = minio_mgr.get_download_url(f["object_name"])

        # 检查是否有音频缺封面，有则后台触发重索引
        try:
            needs_reindex = any(
                not (f.get("file_meta") or {}).get("cover_art")
                for f in audio_files
            )
            if needs_reindex:
                import threading
                threading.Thread(target=minio_mgr.reindex_cover_art, daemon=True).start()
        except Exception:
            pass

        return {"songs": audio_files}
    except Exception as e:
        return {"songs": [], "error": str(e)}

@app.get("/api/music/search")
def search_music(q: str = Query("")):
    """搜索音乐（复用文件管理搜索，过滤只返回音频）"""
    results = minio_mgr.search(q)
    audio_results = [
        f for f in results
        if minio_mgr._is_audio(f.get("content_type", ""), f.get("original_name", ""))
    ]
    for f in audio_results:
        f["download_url"] = minio_mgr.get_download_url(f["object_name"])
    return {"songs": audio_results}

@app.post("/api/music/reindex")
def reindex_music_cover():
    """手动触发封面重新提取"""
    if not minio_mgr.enabled:
        return {"status": "error", "message": "MinIO 未配置"}
    try:
        updated = minio_mgr.reindex_cover_art()
        return {"status": "ok", "updated": updated, "message": f"已更新 {updated} 个文件的封面"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/music/playlist")
def generate_playlist(req: PlaylistRequest):
    """AI 智能生成歌单"""
    if not minio_mgr.enabled:
        return {"status": "error", "message": "MinIO 未配置"}
    try:
        playlist = minio_mgr.generate_playlist(req.prompt)
        return {"status": "ok", "playlist": playlist}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/music/cover/{object_name:path}")
def get_music_cover(object_name: str):
    """获取音频文件的封面图片"""
    import os
    from fastapi import HTTPException  # type: ignore[import]
    from fastapi.responses import FileResponse  # type: ignore[import]

    if not minio_mgr.enabled:
        raise HTTPException(status_code=404, detail="MinIO 未配置")
    
    cover_path = minio_mgr.get_cover_art_file(object_name)
    if cover_path and os.path.exists(cover_path):
        return FileResponse(cover_path)
    
    raise HTTPException(status_code=404, detail="Cover art not found")

@app.get("/api/music/lyrics/{object_name:path}")
def get_music_lyrics(object_name: str):
    """获取音频文件的歌词"""
    from fastapi import HTTPException  # type: ignore[import]
    if not minio_mgr.enabled:
        raise HTTPException(status_code=404, detail="MinIO 未配置")
    
    lyrics = minio_mgr.get_lyrics(object_name)
    return {"lyrics": lyrics}

@app.get("/api/music/stream/{object_name:path}")
def stream_music(object_name: str):
    """获取音频流 URL（重定向到预签名链接）"""
    url = minio_mgr.get_download_url(object_name)
    if not url:
        return {"status": "error", "message": "获取播放链接失败"}
    from fastapi.responses import RedirectResponse  # type: ignore[import]
    return RedirectResponse(url=url)

# ====== WebSocket 端点 ======

@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    ws_clients.add(ws)
    print(f"[WS] 新客户端连接，当前 {len(ws_clients)} 个")
    try:
        while True:
            # 保持连接活跃，同时处理客户端心跳 ping
            data = await ws.receive_text()
            if data == "ping":
                await ws.send_text("pong")
    except WebSocketDisconnect:
        ws_clients.discard(ws)
        print(f"[WS] 客户端断开，剩余 {len(ws_clients)} 个")
    except Exception:
        ws_clients.discard(ws)


# ====== QQ (NapCat) 反向 WebSocket 端点 ======

@app.websocket("/ws/qq")
async def qq_websocket_endpoint(ws: WebSocket):
    """接收 NapCat 反向 WebSocket 推送的 OneBot 11 事件"""
    await ws.accept()
    print("[QQ] NapCat 反向 WebSocket 已连接")

    # 通知守护进程：连接已建立
    if napcat_watchdog:
        napcat_watchdog.report_ws_connect()

    try:
        while True:
            raw = await ws.receive_text()

            # 每收到任意消息都刷新心跳
            if napcat_watchdog:
                napcat_watchdog.report_heartbeat()

            try:
                import json as _json_qq
                event = _json_qq.loads(raw)
            except Exception:
                continue

            # 处理 OneBot 11 事件
            handler = qq_chat_handler
            if handler is not None:
                try:
                    await handler.handle_onebot_event(event)
                except Exception as e:
                    print(f"[QQ] 事件处理异常: {e}")
    except WebSocketDisconnect:
        print("[QQ] NapCat 反向 WebSocket 已断开")
        if napcat_watchdog:
            napcat_watchdog.report_ws_disconnect()
    except Exception as e:
        print(f"[QQ] WebSocket 异常: {e}")
        if napcat_watchdog:
            napcat_watchdog.report_ws_disconnect()


@app.get("/api/qq/status")
def get_qq_status():
    """获取 NapCat / QQ 连接状态"""
    if napcat_watchdog is None:
        return {"status": "disabled", "message": "NapCat 守护未启用"}
    return {"status": "ok", **napcat_watchdog.get_status()}


# ====== 静态文件挂载（放在最后，避免拦截 API 路由） ======
os.makedirs(SESSION_DIR, exist_ok=True)
app.mount("/assets", StaticFiles(directory=SESSION_DIR), name="assets")
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")
